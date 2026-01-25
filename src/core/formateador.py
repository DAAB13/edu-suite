import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

def aplicar_formato_excel(ruta_archivo, nombre_hoja="Sheet1"):
    if not Path(ruta_archivo).exists():
        return

    wb = openpyxl.load_workbook(ruta_archivo)
    ws = wb[nombre_hoja]

    # --- 1. ESTILOS ---
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin', color="D9D9D9"),
        right=Side(style='thin', color="D9D9D9"),
        top=Side(style='thin', color="D9D9D9"),
        bottom=Side(style='thin', color="D9D9D9")
    )

    # --- 2. APLICAR A CABECERA ---
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # --- 3. AUTO-AJUSTE CON LÍMITE MÁXIMO ---
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        # Identificar el nombre de la columna (cabecera)
        header_name = str(column[0].value)

        for cell in column:
            cell.border = thin_border
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # --- LÓGICA DE CONTROL DE ANCHO ---
        # Si la columna es un enlace o grabación, ponemos un límite manual de 30
        # Para el resto, un límite general de 50 para seguridad
        if header_name in ['ENLACE_CLASE', 'GRABACION']:
            adjusted_width = 30
        else:
            adjusted_width = min(max_length + 2, 50)
            
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(ruta_archivo)